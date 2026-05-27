import os
import pandas as pd
from datetime import datetime

# Openpyxl for styled Excel sheets
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab for beautiful PDFs
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# --- EXCEL EXPORTING ---

def export_to_excel(filepath, transactions):
    """
    Exports a list of transactions to a highly styled, professional Excel file.
    """
    if not transactions:
        raise ValueError("No transaction data provided to export.")
        
    wb = Workbook()
    
    # 1. Sheet 1: Transaction Ledger
    ws1 = wb.active
    ws1.title = "Transaction Ledger"
    ws1.views.sheetView[0].showGridLines = True
    
    # Stylings
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E79")
    summary_font = Font(name=font_family, size=11, bold=True)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    income_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    expense_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft orange/red
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Title Block
    ws1["A1"] = "Personal Expense Tracker Ledger"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(name=font_family, size=9, italic=True)
    
    # Ledger Headers
    headers = ["ID", "Date", "Title", "Category", "Type", "Amount", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws1.row_dimensions[4].height = 28
    
    # Fill Data
    row_idx = 5
    for t in transactions:
        ws1.cell(row=row_idx, column=1, value=t.get("id")).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=2, value=t.get("date")).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=3, value=t.get("title"))
        ws1.cell(row=row_idx, column=4, value=t.get("category"))
        
        type_cell = ws1.cell(row=row_idx, column=5, value=t.get("transaction_type"))
        type_cell.alignment = Alignment(horizontal="center")
        if t.get("transaction_type") == "Income":
            type_cell.fill = income_fill
        else:
            type_cell.fill = expense_fill
            
        amt_cell = ws1.cell(row=row_idx, column=6, value=t.get("amount"))
        amt_cell.number_format = '$#,##0.00'
        amt_cell.alignment = Alignment(horizontal="right")
        
        ws1.cell(row=row_idx, column=7, value=t.get("notes") or "")
        
        # Apply zebra formatting and borders
        for col_idx in range(1, 8):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0 and col_idx != 5: # Skip type cell coloring
                cell.fill = zebra_fill
                
        ws1.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Auto-fit Column Widths
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Skip title row from width calculation
        for cell in col:
            if cell.row > 2 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 2. Sheet 2: Executive Dashboard
    ws2 = wb.create_sheet(title="Executive Summary")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "Executive Financial Summary"
    ws2["A1"].font = title_font
    
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 18
    
    summary_headers = ["Key Metrics", "Value"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="333f48", end_color="333f48", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    # Write summary items using formulas pointing to ledger sheet
    ledger_rows = len(transactions)
    
    ws2.cell(row=4, column=1, value="Total Income").font = data_font
    income_cell = ws2.cell(row=4, column=2, value=f"=SUMIF('Transaction Ledger'!E5:E{ledger_rows+4}, \"Income\", 'Transaction Ledger'!F5:F{ledger_rows+4})")
    income_cell.font = data_font
    income_cell.number_format = '$#,##0.00'
    
    ws2.cell(row=5, column=1, value="Total Expenses").font = data_font
    expense_cell = ws2.cell(row=5, column=2, value=f"=SUMIF('Transaction Ledger'!E5:E{ledger_rows+4}, \"Expense\", 'Transaction Ledger'!F5:F{ledger_rows+4})")
    expense_cell.font = data_font
    expense_cell.number_format = '$#,##0.00'
    
    ws2.cell(row=6, column=1, value="Net Savings").font = summary_font
    net_cell = ws2.cell(row=6, column=2, value="=B4-B5")
    net_cell.font = summary_font
    net_cell.number_format = '$#,##0.00'
    net_cell.border = double_bottom_border
    
    # Save spreadsheet
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)


# --- PDF REPORTING ---

class NumberedCanvas(canvas.Canvas):
    """
    Custom canvas to calculate total page count dynamically 
    and draw 'Page X of Y' on every footer.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#555555"))
        
        # Draw header running rule
        self.setStrokeColor(colors.HexColor("#D9D9D9"))
        self.setLineWidth(0.5)
        self.line(54, 750, 558, 750)
        self.drawString(54, 755, "Personal Expense Tracker - Financial Report")
        
        # Draw footer
        self.line(54, 50, 558, 50)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 38, page_text)
        self.drawString(54, 38, f"Report Generated: {datetime.now().strftime('%Y-%m-%d')}")
        self.restoreState()


def export_to_pdf(filepath, transactions, summary_stats):
    """
    Generates a beautifully typeset, executive PDF report using ReportLab.
    """
    # Page settings: Margins are 0.75 in (54 pt)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    
    # Create custom beautiful styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor("#555555"),
        spaceAfter=25
    )
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor("#333333"),
        spaceBefore=15,
        spaceAfter=10
    )
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#222222")
    )
    table_header_style = ParagraphStyle(
        'TableHeaderText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1 # centered
    )
    
    story = []
    
    # 1. Header Banner Block
    story.append(Paragraph("FINANCIAL PERFORMANCE REPORT", title_style))
    story.append(Paragraph(f"Analysis Period: Current Month Ledger | Extracted on {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # 2. KPI Cards Grid
    kpi_data = [
        [
            Paragraph("<b>Total Income</b>", table_text_style), 
            Paragraph("<b>Total Expenses</b>", table_text_style), 
            Paragraph("<b>Net Savings</b>", table_text_style), 
            Paragraph("<b>Savings Rate</b>", table_text_style)
        ],
        [
            f"${summary_stats['total_income']:,.2f}", 
            f"${summary_stats['total_expense']:,.2f}", 
            f"${summary_stats['net_savings']:,.2f}", 
            f"{summary_stats['savings_rate']:.1f}%"
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[126, 126, 126, 126])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EEF3F7")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,1), (-1,1), 8),
        ('TOPPADDING', (0,1), (-1,1), 8),
        ('TEXTCOLOR', (0,1), (0,1), colors.HexColor("#27AE60")), # Income Green
        ('TEXTCOLOR', (1,1), (1,1), colors.HexColor("#C0392B")), # Expense Red
        ('TEXTCOLOR', (2,1), (2,1), colors.HexColor("#1F4E79")), # Savings Blue
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 12),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#D9D9D9")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E2E2")),
    ]))
    
    story.append(Paragraph("Key Financial Indicators", h2_style))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # 3. Transaction Log Table
    story.append(Paragraph("Detailed Transactions Ledger", h2_style))
    
    # Define widths adding up to 504 pt (which matches letter width 612 - 108 margin)
    col_widths = [60, 134, 100, 70, 70, 70] 
    
    # Table headers
    th = [
        Paragraph("Date", table_header_style), 
        Paragraph("Title", table_header_style), 
        Paragraph("Category", table_header_style), 
        Paragraph("Type", table_header_style), 
        Paragraph("Amount", table_header_style),
        Paragraph("Flow", table_header_style)
    ]
    
    table_data = [th]
    
    # Populate transactions
    for t in transactions:
        is_inc = t.get("transaction_type") == "Income"
        flow_indicator = "+" if is_inc else "-"
        amt_str = f"${t.get('amount'):,.2f}"
        
        row = [
            Paragraph(t.get("date"), table_text_style),
            Paragraph(t.get("title"), table_text_style),
            Paragraph(t.get("category"), table_text_style),
            Paragraph(t.get("transaction_type"), table_text_style),
            Paragraph(amt_str, table_text_style),
            Paragraph(flow_indicator, ParagraphStyle('FlowStyle', parent=table_text_style, fontName='Helvetica-Bold', textColor=colors.HexColor("#27AE60" if is_inc else "#C0392B"), alignment=1))
        ]
        table_data.append(row)
        
    trans_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Table styling
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E2E2")),
    ]
    
    # Add zebra stripes
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor("#F9FBFC")))
            
    trans_table.setStyle(TableStyle(t_style))
    story.append(trans_table)
    
    # Build the document
    doc.build(story, canvasmaker=NumberedCanvas)


# --- CSV IMPORTING ---

def import_from_csv(filepath):
    """
    Imports and validates transactions from a CSV file.
    Expected headers: title, amount, category, transaction_type, date, (optional: notes)
    Returns:
      list of dict: list of validated, parsed transactions.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Selected file path does not exist: {filepath}")
        
    try:
        df = pd.read_csv(filepath, encoding="utf-8-sig")
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(filepath, encoding="latin-1")
        except Exception as e:
            raise ValueError(f"Could not read CSV (encoding issue): {e}")
    except Exception as e:
        raise ValueError(f"Failed to read CSV: {e}")

    # Normalize headers: lowercase + replace spaces/hyphens with underscores
    df.columns = [c.strip().lower().replace(" ", "_").replace("-", "_") for c in df.columns]
    required_cols = ["title", "amount", "category", "transaction_type", "date"]
    
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns in CSV: {', '.join(missing_cols)}.\nHeaders must be: title, amount, category, transaction_type, date")
        
    parsed_transactions = []
    
    for idx, row in df.iterrows():
        line_num = idx + 2 # 1-based index + header row offset
        
        # Title validation
        title = str(row["title"]).strip()
        if not title or title.lower() == "nan":
            raise ValueError(f"Row {line_num}: 'title' cannot be empty.")
            
        # Amount validation
        try:
            amount = float(row["amount"])
            if amount <= 0:
                raise ValueError()
        except Exception:
            raise ValueError(f"Row {line_num}: 'amount' must be a positive number. Got: {row['amount']}")
            
        # Category validation
        category = str(row["category"]).strip()
        if not category or category.lower() == "nan":
            category = "Other"
            
        # Transaction Type validation
        tx_type = str(row["transaction_type"]).strip().title()
        if tx_type not in ["Income", "Expense"]:
            raise ValueError(f"Row {line_num}: 'transaction_type' must be 'Income' or 'Expense'. Got: {row['transaction_type']}")
            
        # Date validation
        date_str = str(row["date"]).strip()
        try:
            # Validate ISO format YYYY-MM-DD
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            # Try parsing from common formats and convert to YYYY-MM-DD
            parsed_ok = False
            for fmt in ["%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    date_str = dt.strftime("%Y-%m-%d")
                    parsed_ok = True
                    break
                except ValueError:
                    continue
            if not parsed_ok:
                raise ValueError(f"Row {line_num}: 'date' must be in YYYY-MM-DD format (or common variations). Got: {row['date']}")
                
        # Notes (optional)
        notes = ""
        if "notes" in df.columns:
            val = str(row["notes"]).strip()
            notes = "" if val.lower() == "nan" else val
            
        parsed_transactions.append({
            "title": title,
            "amount": amount,
            "category": category,
            "transaction_type": tx_type,
            "date": date_str,
            "notes": notes
        })
        
    return parsed_transactions
